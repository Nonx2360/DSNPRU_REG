import os
import re
import hashlib
from io import BytesIO
from datetime import datetime
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from .. import models
from ..auth import get_current_admin
from ..database import get_db

router = APIRouter()


def _sanitize_filename(name: str) -> str:
    name = re.sub(r'[\\/*?:"<>|]', "", name or "")
    return name.strip()[:200] or "registrations"


def _get_registrations(db: Session, activity_id: Optional[int] = None):
    query = db.query(models.Registration).join(models.Student).join(models.Activity)
    if activity_id:
        query = query.filter(models.Registration.activity_id == activity_id)
    return query.all()


def _pastel_hex(seed: str) -> str:
    h = int(hashlib.md5(seed.encode()).hexdigest(), 16)
    red = 190 + ((h & 0xFF) % 50)
    green = 190 + (((h >> 8) & 0xFF) % 50)
    blue = 190 + (((h >> 16) & 0xFF) % 50)
    return f"{red:02X}{green:02X}{blue:02X}"


def _safe_sheet_title(name: str) -> str:
    name = re.sub(r"[\[\]:*?/\\]", "", name or "")
    return (name.strip() or "Registrations")[:31]


def _style_registration_sheet(ws, is_team: bool, regs, activity_title: str = ""):
    thin_gray = Side(style="thin", color="D1D5DB")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="E5E7EB")
    unnamed_team_fill = PatternFill("solid", fgColor="FFE666")

    max_column = ws.max_column
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    ws.cell(1, 1).value = "รายชื่อผู้ลงทะเบียน"
    if activity_title:
        ws.cell(1, 1).value = f"รายชื่อผู้ลงทะเบียน: {activity_title}"
    ws.cell(1, 1).fill = title_fill
    ws.cell(1, 1).font = Font(bold=True, size=14, color="111827")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 24

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row[0].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if is_team:
            team_name = row[1].value
            row[1].fill = unnamed_team_fill if team_name == "-" else PatternFill("solid", fgColor=_pastel_hex(str(team_name)))
            row[1].font = Font(bold=True, color="111827")

    widths = [34, 22, 30, 14, 16] if is_team else [38, 32, 14, 16]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_column)}{ws.max_row}"


def _add_registration_sheet(wb, activity, regs, title=None):
    is_team = activity.type == "team"
    if is_team:
        regs = sorted(regs, key=lambda x: (x.team_name or "", x.student.classroom or "", x.student.sequence or 0))
    else:
        regs = sorted(regs, key=lambda x: (x.student.classroom or "", x.student.sequence or 0))

    ws = wb.create_sheet(_safe_sheet_title(title or activity.title))
    if is_team:
        ws.append(["กิจกรรม", "ทีม", "ชื่อ-สกุล", "ห้อง", "รหัสนักเรียน"])
        for r in regs:
            ws.append([
                r.activity.title,
                r.team_name or "-",
                r.student.name,
                r.student.classroom,
                r.student.number,
            ])
    else:
        ws.append(["กิจกรรม", "ชื่อ-สกุล", "ห้อง", "รหัสนักเรียน"])
        for r in regs:
            ws.append([
                r.activity.title,
                r.student.name,
                r.student.classroom,
                r.student.number,
            ])

    _style_registration_sheet(ws, is_team, regs, activity.title)
    return ws


@router.get("/excel")
def export_excel(
    activity_id: Optional[int] = None,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    wb = Workbook()
    wb.remove(wb.active)

    if activity_id:
        activity = db.query(models.Activity).filter(models.Activity.id == activity_id).first()
        regs = _get_registrations(db, activity_id)
        if activity:
            _add_registration_sheet(wb, activity, regs)
            filename = f"{_sanitize_filename(activity.title)}.xlsx"
        elif regs:
            _add_registration_sheet(wb, regs[0].activity, regs)
            filename = f"{_sanitize_filename(regs[0].activity.title)}.xlsx"
        else:
            ws = wb.create_sheet("Registrations")
            ws.append(["กิจกรรม", "ชื่อ-สกุล", "ห้อง", "รหัสนักเรียน"])
            _style_registration_sheet(ws, False, [], "")
            filename = "registrations.xlsx"
    else:
        activities = db.query(models.Activity).order_by(models.Activity.id).all()
        all_regs = _get_registrations(db)
        regs_by_activity = {}
        for reg in all_regs:
            regs_by_activity.setdefault(reg.activity_id, []).append(reg)
        if activities:
            used_titles = {}
            for activity in activities:
                base_title = _safe_sheet_title(activity.title)
                used_titles[base_title] = used_titles.get(base_title, 0) + 1
                if used_titles[base_title] == 1:
                    sheet_title = base_title
                else:
                    suffix = f"_{used_titles[base_title]}"
                    sheet_title = f"{base_title[:31 - len(suffix)]}{suffix}"
                _add_registration_sheet(wb, activity, regs_by_activity.get(activity.id, []), sheet_title)
        else:
            ws = wb.create_sheet("Registrations")
            ws.append(["กิจกรรม", "ชื่อ-สกุล", "ห้อง", "รหัสนักเรียน"])
            _style_registration_sheet(ws, False, [], "")
        filename = f"all_activity_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    ascii_filename = filename.encode("ascii", "ignore").decode() or "registrations.xlsx"
    return Response(
        content=stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_filename}"; '
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )


@router.get("/pdf")
def export_pdf(
    activity_id: Optional[int] = None,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    regs = _get_registrations(db, activity_id)
    stream = BytesIO()

    # Register SukhumvitSet font for Thai text support
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fonts_dir = os.path.join(base_dir, "frontend", "static", "fonts")
    font_name = "ChakraPetch"
    
    try:
        font_weights = ["Regular", "Medium", "SemiBold", "Bold"]
        font_path = None
        for weight in font_weights:
            test_path = os.path.join(fonts_dir, f"ChakraPetch-{weight}.ttf")
            if os.path.exists(test_path):
                font_path = test_path
                break
        
        if font_path:
            pdfmetrics.registerFont(TTFont("ChakraPetch", font_path))
        else:
            font_name = "Helvetica"
    except Exception:
        font_name = "Helvetica"

    doc = SimpleDocTemplate(stream, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    thai_style = ParagraphStyle(
        'ThaiStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=16,
        leading=20,
        alignment=1, # Center
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=11,
        leading=14,
        alignment=1, # Center
    )
    
    title_text = "รายชื่อผู้ลงทะเบียนกิจกรรม DSNPRU_REG"
    if activity_id and regs:
        title_text = f"รายชื่อผู้ลงทะเบียน: {regs[0].activity.title}"
    
    elements.append(Paragraph(title_text, thai_style))
    elements.append(Spacer(1, 20))

    is_team = False
    if regs and hasattr(regs[0].activity, 'type') and regs[0].activity.type == "team":
        is_team = True
        regs = sorted(regs, key=lambda x: (x.team_name or "", x.student.classroom or "", x.student.sequence or 0))
    else:
        regs = sorted(regs, key=lambda x: (x.student.classroom or "", x.student.sequence or 0))

    def get_team_color(seed, dark=False):
        import hashlib
        h = int(hashlib.md5(str(seed).encode()).hexdigest(), 16)
        if dark:
            r = (h & 0xFF) % 150 # darker colors for white text
            g = ((h >> 8) & 0xFF) % 150
            b = ((h >> 16) & 0xFF) % 150
        else:
            r = 180 + ((h & 0xFF) % 65)
            g = 180 + (((h >> 8) & 0xFF) % 65)
            b = 180 + (((h >> 16) & 0xFF) % 65)
        return colors.Color(r/255.0, g/255.0, b/255.0)

    unnamed_team_color = colors.Color(1.0, 0.90, 0.35)

    def get_team_paragraph(name):
        if not name:
            style = ParagraphStyle(
                'UnnamedTeam',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=11,
                alignment=1,
                textColor=colors.black,
            )
            return Paragraph("-", style)

        bg_color = get_team_color(name, dark=True)
        import hashlib
        h = int(hashlib.md5(name.encode()).hexdigest(), 16)
        
        style = ParagraphStyle(
            f'Team_{h}',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=11,
            alignment=1,
            textColor=colors.white,
            backColor=bg_color,
            borderPadding=4
        )
        return Paragraph(f"{name}", style)

    if is_team:
        data = [["ลำดับ", "กิจกรรม", "ชื่อทีม", "ชื่อ-สกุล", "ห้อง", "รหัส"]]
        for i, r in enumerate(regs, 1):
            data.append([
                str(i),
                Paragraph(r.activity.title or "-", cell_style),
                get_team_paragraph(r.team_name),
                Paragraph(r.student.name or "-", cell_style),
                r.student.classroom or "-",
                r.student.number
            ])
        colWidths = [40, 100, 100, 140, 70, 60]
    else:
        data = [["ลำดับ", "กิจกรรม", "ชื่อ-สกุล", "ห้อง", "รหัสนักเรียน"]]
        for i, r in enumerate(regs, 1):
            data.append([
                str(i),
                Paragraph(r.activity.title or "-", cell_style),
                Paragraph(r.student.name or "-", cell_style),
                r.student.classroom or "-",
                r.student.number
            ])
        colWidths = [40, 140, 160, 80, 90]

    table_style_commands = [
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.rose if hasattr(colors, 'rose') else colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]
    if is_team:
        for row_index, reg in enumerate(regs, 1):
            if not reg.team_name:
                table_style_commands.append(
                    ('BACKGROUND', (2, row_index), (2, row_index), unnamed_team_color)
                )

    t = Table(data, colWidths=colWidths)
    t.setStyle(TableStyle(table_style_commands))
    
    elements.append(t)
    
    def add_footer(canvas, doc):
        canvas.saveState()
        try:
            canvas.setFont(font_name, 9)
        except:
            canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(A4[0] - 30, 15, "Auto Generate Using DSNPRU_REG By ณัฐชนน รอดน้อย 04824")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    
    stream.seek(0)
    filename = f"registrations_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    
    return Response(
        content=stream.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
@router.get("/students/excel")
def export_students_excel(
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    students = db.query(models.Student).order_by(models.Student.classroom, models.Student.sequence).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["รหัส", "คำนำหน้า", "ชื่อ", "นามสกุล", "ห้อง", "เลขที่"])

    prefixes = ["เด็กชาย", "เด็กหญิง", "นาย", "นางสาว", "ด.ช.", "ด.ญ."]
    
    for s in students:
        full_name = s.name.strip()
        found_prefix = ""
        rest_of_name = full_name
        
        # Try to extract prefix
        for p in prefixes:
            if full_name.startswith(p):
                found_prefix = p
                rest_of_name = full_name[len(p):].strip()
                break
        
        # Split first and last name
        name_parts = rest_of_name.split(" ", 1)
        first_name = name_parts[0] if name_parts else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""
        
        ws.append([s.number, found_prefix, first_name, last_name, s.classroom or "", s.sequence or ""])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/students/pdf")
def export_students_pdf(
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    students = db.query(models.Student).order_by(models.Student.classroom, models.Student.sequence).all()
    stream = BytesIO()

    # Register SukhumvitSet font for Thai text support
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fonts_dir = os.path.join(base_dir, "frontend", "static", "fonts")
    font_name = "ChakraPetch"
    
    try:
        font_weights = ["Regular", "Medium", "SemiBold", "Bold"]
        font_path = None
        for weight in font_weights:
            test_path = os.path.join(fonts_dir, f"ChakraPetch-{weight}.ttf")
            if os.path.exists(test_path):
                font_path = test_path
                break
        
        if font_path:
            pdfmetrics.registerFont(TTFont("ChakraPetch", font_path))
        else:
            font_name = "Helvetica"
    except Exception:
        font_name = "Helvetica"

    doc = SimpleDocTemplate(stream, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    # Custom Thai Paragraph Style
    styles = getSampleStyleSheet()
    thai_style = ParagraphStyle(
        'ThaiStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=16,
        leading=20,
        alignment=1, # Center
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=11,
        leading=14,
        alignment=1, # Center
    )
    
    elements.append(Paragraph("รายชื่อนักเรียนทั้งหมด - DSNPRU_REG", thai_style))
    elements.append(Spacer(1, 20))

    # Table Data
    data = [["ลำดับ", "รหัสนักเรียน", "ชื่อ-นามสกุล", "ห้อง", "เลขที่"]]
    for i, s in enumerate(students, 1):
        data.append([
            str(i),
            s.number,
            Paragraph(s.name or "-", cell_style),
            s.classroom or "-",
            str(s.sequence) if s.sequence else "-"
        ])

    # Table Styling
    t = Table(data, colWidths=[40, 90, 240, 90, 50])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.rose if hasattr(colors, 'rose') else colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(t)
    
    def add_footer(canvas, doc):
        canvas.saveState()
        try:
            canvas.setFont(font_name, 9)
        except:
            canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(A4[0] - 30, 15, "Auto Generate Using DSNPRU_REG By ณัฐชนน รอดน้อย 04824")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    
    stream.seek(0)
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    
    return Response(
        content=stream.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
