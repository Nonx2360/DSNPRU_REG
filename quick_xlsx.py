import hashlib
import os
import re
import sys
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
from sqlalchemy.orm import joinedload, sessionmaker

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from backend.models import Activity, Registration

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "sicday.db")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "quick_export")

engine = create_engine(f"sqlite:///{DB_PATH}", connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)


def sanitize_filename(name):
    name = re.sub(r'[\\/*?:"<>|]', "", name or "")
    return name.strip()[:200] or "unnamed"


def pastel_hex(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest(), 16)
    red = 190 + ((h & 0xFF) % 50)
    green = 190 + (((h >> 8) & 0xFF) % 50)
    blue = 190 + (((h >> 16) & 0xFF) % 50)
    return f"{red:02X}{green:02X}{blue:02X}"


def style_sheet(ws, activity, is_team):
    thin_gray = Side(style="thin", color="D1D5DB")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="1F2937")
    title_fill = PatternFill("solid", fgColor="E5E7EB")
    unnamed_team_fill = PatternFill("solid", fgColor="FFE666")

    max_column = ws.max_column
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    ws.cell(1, 1).value = f"รายชื่อผู้ลงทะเบียน: {activity.title}"
    ws.cell(1, 1).fill = title_fill
    ws.cell(1, 1).font = Font(bold=True, size=14, color="111827")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 24

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row[0].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if is_team:
            team_name = row[1].value
            row[1].fill = unnamed_team_fill if team_name == "-" else PatternFill("solid", fgColor=pastel_hex(str(team_name)))
            row[1].font = Font(bold=True, color="111827")

    widths = [34, 22, 30, 14, 16] if is_team else [38, 32, 14, 16]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_column)}{ws.max_row}"


def build_xlsx(activity, regs):
    is_team = activity.type == "team"
    if is_team:
        regs = sorted(regs, key=lambda x: (x.team_name or "", x.student.classroom or "", x.student.sequence or 0))
    else:
        regs = sorted(regs, key=lambda x: (x.student.classroom or "", x.student.sequence or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    if is_team:
        ws.append(["กิจกรรม", "ทีม", "ชื่อ-สกุล", "ห้อง", "รหัสนักเรียน"])
        for reg in regs:
            ws.append([
                reg.activity.title,
                reg.team_name or "-",
                reg.student.name,
                reg.student.classroom,
                reg.student.number,
            ])
    else:
        ws.append(["กิจกรรม", "ชื่อ-สกุล", "ห้อง", "รหัสนักเรียน"])
        for reg in regs:
            ws.append([
                reg.activity.title,
                reg.student.name,
                reg.student.classroom,
                reg.student.number,
            ])

    style_sheet(ws, activity, is_team)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    db = SessionLocal()

    try:
        activities = db.query(Activity).all()
        if not activities:
            print("No activities found.")
            return

        print(f"Found {len(activities)} activities. Exporting XLSX files...\n")

        for activity in activities:
            regs = db.query(Registration).options(
                joinedload(Registration.student),
                joinedload(Registration.activity),
            ).filter(Registration.activity_id == activity.id).all()

            filename = sanitize_filename(activity.title) + ".xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)
            xlsx_stream = build_xlsx(activity, regs)
            with open(filepath, "wb") as file:
                file.write(xlsx_stream.read())

            print(f"  {filename}  ({len(regs)} registrations)")

        print(f"\nDone! XLSX files saved to: {OUTPUT_DIR}")
    finally:
        db.close()


if __name__ == "__main__":
    import io

    if sys.platform == "win32":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
